import requests
import pandas as pd
from openpyxl import load_workbook
import json

# Load the configuration file
with open('config3.json', 'r') as file:
    config = json.load(file)

#set variables
year = '2021'
dsource = 'acs'
dname = 'acs5'
cols = 'group(S1810)'
state = '45' #Don't change this


# 001 - Abbeville, 003 - Aiken, 005 - Allendale, 007 - Anderson, 009 - Bamberg, 011 - Barnwell, 013 - Beaufort, 015 - Berkeley, 017 - Calhoun, 019 - Charleston, 
# 021 - Cherokee, 023 - Chester, 025 - Chesterfield, 027 - Clarendon, 029 - Colleton, 031 - Darlington, 033 - Dillon, 035 - Dorchester, 037 - Edgefield, 039 - Fairfield,
# 041 - Florence, 043 - Georgetown, 045 - Greenville, 047 - Greenwood, 049 - Hampton, 051 - Horry, 053 - Jasper, 055 - Kershaw, 057 - Lancaster, 059 - Laurens, 061 - Lee, 
# 063 - Lexington, 065 - McCormick, 067 - Marion, 069 - Marlboro, 071 - Newberry, 073 - Oconee, 075 - Orangeburg, 077 - Pickens, 079 - Richland, 081 - Saluda, 
# 083 - Spartanburg, 085 - Sumter, 087 - Union, 089 - Williamsburg, 091 - York


county = '*'
keyfile = config['keyfile'] # Get the keyfile path from the configuration file

#Change outfile to which ever file you would like to save data to
outfile = config['outfile'] # Get the outfile path from the configuration file

#If a sheet name that does not exist is entered, one will be created with that name
sheet_name = 'Sheet1'

s_path = 'subject' # set to either 'subject' or 'spp' for tables that begin with 'S'

#construct base_url
if cols.startswith('group(S'):
    base_url = f'https://api.census.gov/data/{year}/{dsource}/{dname}/{s_path}'
    # If the code does not run correctly, try changing the value of s_path to 'spp'
elif cols.startswith('group(CP'):
    base_url = f'https://api.census.gov/data/{year}/{dsource}/{dname}/cprofile'
else:
    base_url = f'https://api.census.gov/data/{year}/{dsource}/{dname}'

#read api key in from file
with open(keyfile) as key:
    api_key = key.read().strip()

#retrieve data, print output to screen
data_url = f'{base_url}?get={cols}&for=county:{county}&in=state:{state}&key={api_key}'
response = requests.get(data_url, verify = False)
popdata = response.json()

# Load the CSV file into a DataFrame
df = pd.read_csv(config['csv_file_path'])

# Extract the first column of the DataFrame, including the column header
labels = df.iloc[:,0].tolist()

# Replace non-breaking space characters with regular space characters
labels = [label.replace('\xa0', '').replace(' ', '') for label in labels]

# Create a pandas DataFrame from the data
headers = popdata.pop(0)
new_df = pd.DataFrame(popdata, columns=headers)

# Filter the DataFrame to exclude columns that end in "M", "EA", or "MA"
filtered_columns = [col for col in new_df.columns if not col.endswith('M') and not col.endswith('EA') and not col.endswith('MA')]
new_df = new_df[filtered_columns]

#Makes  labels the new column headers
new_df.columns = labels + [col for col in new_df.columns[-4:]]

# Replace all None values in new_df with zeros
new_df = new_df.fillna(0)

# Replace all instances of -888888888 and -666666666.0 in new_df with 'n/a'
new_df = new_df.replace(['-888888888', '-666666666.0'], 'n/a')

# Extract county names and Total civilian noninstitutionalized population numbers from new_df
county_names = new_df['NAME']
population_numbers = new_df['S1810_C01_001E']

# Keep only the part of each county name before the first comma
county_names = county_names.str.split(',').str[0]

# Create a new DataFrame with the desired format
data = {'Year': [year]}
for county, population in zip(county_names, population_numbers):
    data[county] = [population]
new_df = pd.DataFrame(data)
 
# Sort columns alphabetically, excluding the "Year" column
new_df = new_df[['Year'] + sorted([col for col in new_df.columns if col != 'Year'])]

# Load existing Excel file and read data from specified sheet into a DataFrame
book = load_workbook(outfile)
if sheet_name in book.sheetnames:
    ws = book[sheet_name]
    data_rows = []
    for row in ws.iter_rows(values_only=True):
        data_rows.append(row)
    if data_rows:
        old_df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
    else:
        old_df = pd.DataFrame()
else:
    old_df = pd.DataFrame()
    


# Combine old and new data
combined_df = pd.concat([old_df, new_df], axis=0, ignore_index=True)

# Write combined data to Excel file
from openpyxl.utils.dataframe import dataframe_to_rows

book = load_workbook(outfile) # Load the existing Excel file
if sheet_name in book.sheetnames: # Check if the specified sheet name exists in the workbook
    ws = book[sheet_name] # Get the specified worksheet
    for row in ws.iter_rows(): # Iterate over each row in the worksheet
        for cell in row:
            cell.value = None # Clear the cell value
else:
    ws = book.create_sheet(sheet_name) # Create a new worksheet with the specified name

# Convert the combined DataFrame to rows and iterate over each row
for r_index, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True)):
    for c_index, value in enumerate(row):
        ws.cell(row=r_index+1, column=c_index+1, value=value) # Write the value to the corresponding cell in the worksheet

book.save(outfile) # Save the changes to the Excel file